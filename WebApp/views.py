# Importaciones de Python estándar
from datetime import datetime, timedelta, time, date

# Configuraciones de Django
from django.conf import settings

# Autenticación y autorización de Django
from django.contrib import auth
from django.contrib.auth.decorators import login_required
from django.contrib.auth.tokens import default_token_generator
from django.contrib import messages  # Para mostrar mensajes de éxito, error, etc.

# Envío de correo electrónico
from django.core.mail import send_mail

# Funciones y clases de vista de Django
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.template.loader import render_to_string
from django.urls import reverse

# Funciones y clases para consultas de base de datos Django
from django.db.models import Q
from django.db import IntegrityError
from .forms import *
from user.models import *

# Manipulación de archivos Excel
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

# Generación de archivos PDF
from reportlab.lib.pagesizes import letter, landscape
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from io import BytesIO
from reportlab.lib.units import mm


#Vista home
def home(request):
    return render(request, 'home.html')

#Vista 404
def paginaNoEncontrada(request, exception):
    return render(request, 'paginaNoEncontrada.html', status=404)

# Vista Login
def login(request):
    if request.method == "POST":
        email = request.POST.get("email")
        password = request.POST.get("password")

        user = auth.authenticate(email=email, password=password)

        if user is not None:
            auth.login(request, user)
            messages.success(request, "Bienvenido " + user.nombre + " " + user.apellido)
            return redirect("perfil") 
        else:
            messages.error(request, "Credenciales incorrectas")

    return render(request, 'registration/login.html')

#Logout
@login_required
def logout(request):
    auth.logout(request)
    return redirect("home")

#Vista Perfil
@login_required
def perfil(request):
    if request.user.tipoUsuario.nombre_tipo_usuario in ["Fonoaudiologo", "Gerencia", "Tutor"]:
        today = date.today()
        tomorrow = today + timedelta(days=1)

        reservas_hoy = ReservaHora.objects.filter(fecha=today, estado='Reservada').order_by('id')
        reservas_futuras = ReservaHora.objects.filter(fecha__gte=tomorrow, estado='Reservada').order_by('id')
        
        data = {
            'reservas_hoy': reservas_hoy,
            'reservas_total': ReservaHora.objects.exclude(estado='Reservada').order_by('-id'),
            'reservas_futuras': reservas_futuras,
            'logs': Log.objects.all().order_by('-id')
        }
        return render(request, 'perfil.html', data)
    else:
        return render(request, 'perfil.html')

#Vista Equipo
def equipo(request):
    return render(request, 'equipo.html')

#Vista Nosotros
def nosotros(request):
    return render(request, 'nosotros.html')


#Oirs
def oirs(request):
    if request.user.is_authenticated and request.user.tipoUsuario.nombre_tipo_usuario == 'Gerencia':
        oirs_data = OIRS.objects.all()
        return render(request, 'oirs/oirs.html', {'oirs_data': oirs_data})
    elif request.method == 'POST':
        form = OIRSMensajeForm(request.POST)
        if form.is_valid():
            form.save()
            
            # Extraer los datos del formulario
            tipo_mensaje = form.cleaned_data.get('tipo_mensaje')
            nombre = form.cleaned_data.get('nombre')
            email = form.cleaned_data.get('email')
            mensaje = form.cleaned_data.get('mensaje')
            
            log = Log(username = email, texto = 'Envío de Mensaje OIRS')
            log.save()
            
            # Renderizar el template del correo electrónico
            html_message = render_to_string('oirs/correoOirs.html', {
                'tipo_mensaje': tipo_mensaje,
                'nombre': nombre,
                'email': email,
                'mensaje': mensaje,
            })

            # Preparar el contenido del correo electrónico
            subject = 'Nuevo mensaje OIRS'
            from_email = settings.EMAIL_HOST_USER
            recipient_list = ['correo.proyectos.duoc@gmail.com']  # Dirección de correo a la que se enviará el mensaje

            # Intentar enviar el correo electrónico
            try:
                send_mail(subject, None, from_email, recipient_list, html_message=html_message)
                messages.success(request, 'Su mensaje ha sido enviado con éxito.')
                return redirect('oirs')
            except Exception as e:
                messages.error(request, 'Error al enviar el correo electrónico. Por favor, inténtelo de nuevo más tarde.')
    else:
        form = OIRSMensajeForm()
    return render(request, 'oirs/oirs.html', {'form': form})

# ------------------- Reserva de horas -------------------

#Calendario
def calendario(request):
    context = {
        'doctores': Fonoaudiologo.objects.all(),
    }
    return render(request, 'reservaHoras/calendario.html', context)

# Horas Disponibles
def ver_horas_disponibles(request):
    fecha_reserva_str = request.GET.get('fecha_reserva')
    doctor_id = request.GET.get('doctor_id')

    if fecha_reserva_str and doctor_id:
        fecha_reserva = datetime.strptime(fecha_reserva_str, '%d-%m-%Y')
        fecha_reserva_str_dd_mm_yyyy = fecha_reserva.strftime('%d/%m/%Y')
        doctor = Fonoaudiologo.objects.get(pk=doctor_id)
        horas_disponibles = obtener_horas_disponibles_para_doctor(doctor, fecha_reserva)
        context = {
            'fecha_reserva': fecha_reserva_str,
            'fecha_reserva_dt': fecha_reserva_str_dd_mm_yyyy,
            'doctor': doctor,
            'horas_disponibles': horas_disponibles
        }
        return render(request, 'reservaHoras/horasDisponibles.html', context)
    else:
        messages.error(request, 'Debe seleccionar una fecha y un doctor para ver las horas disponibles')
        return redirect('calendario')

def obtener_horas_disponibles_para_doctor(doctor, fecha_reserva):
    dia_semana = fecha_reserva.weekday()
    horas_trabajo = HorasTrabajo.objects.filter(doctor=doctor, dia_semana=dia_semana)

    if horas_trabajo.exists():
        hora_inicio = horas_trabajo.first().hora_inicio
        hora_fin = horas_trabajo.first().hora_fin
    else:
        return []
    
    ahora = datetime.now().time()
    
    fecha_inicio = datetime.combine(fecha_reserva, time())
    hora_inicio_dt = fecha_inicio.replace(hour=hora_inicio.hour, minute=hora_inicio.minute)
    hora_fin_dt = fecha_inicio.replace(hour=hora_fin.hour, minute=hora_fin.minute)

    todas_las_horas = []
    hora_actual = hora_inicio_dt
    while hora_actual < hora_fin_dt:
        if fecha_reserva.date() > datetime.now().date() or (fecha_reserva.date() == datetime.now().date() and hora_actual.time() > ahora):
            todas_las_horas.append(hora_actual.strftime('%H:%M'))
        hora_actual += timedelta(hours=1)

    citas = ReservaHora.objects.filter(fonoaudiologo=doctor, fecha=fecha_reserva)
    horas_reservadas = [cita.hora.strftime('%H:%M') for cita in citas]
    horas_disponibles = [hora for hora in todas_las_horas if hora not in horas_reservadas]

    return horas_disponibles


#Formulario de Reserva
def reservaHora(request):
    fecha_reserva = request.GET.get('fecha')
    doctor_id = request.GET.get('fonoaudiologo')
    hora = request.GET.get('hora')
    doctor = Fonoaudiologo.objects.get(pk=doctor_id)
    
    if request.method == 'GET' and request.user.is_authenticated:
        if request.user.tipoUsuario.nombre_tipo_usuario == 'Tutor':
            tutor = Tutor.objects.get(emailTutor=request.user.email)
            paciente = Paciente.objects.get(tutor=tutor)
            data = {
                'form': ReservaHoraForm(initial={'nombrePaciente': paciente.nombre, 'apellidoPaciente': paciente.apellido, 'rutPaciente': paciente.rut, 'telefonoPaciente': paciente.telefono, 'emailPaciente': paciente.tutor.emailTutor}),
                'fecha_reserva': fecha_reserva,
                'hora': hora,
                'doctor': doctor
            }
        else:   
            data = {
                'form': ReservaHoraForm(),
                'fecha_reserva': fecha_reserva,
                'hora': hora,
                'doctor': doctor
            }
    else:
        data = {
            'form': ReservaHoraForm(),
            'fecha_reserva': fecha_reserva,
            'hora': hora,
            'doctor': doctor
        }
        
    if request.method == 'POST':
        formulario = ReservaHoraForm(request.POST)
        if formulario.is_valid():
            nombre = formulario.cleaned_data.get('nombrePaciente')
            apellido = formulario.cleaned_data.get('apellidoPaciente')
            rut = formulario.cleaned_data.get('rutPaciente')
            telefono = formulario.cleaned_data.get('telefonoPaciente')
            email = formulario.cleaned_data.get('emailPaciente')
            
            res = ReservaHora()
            res.fecha = datetime.strptime(fecha_reserva, '%d-%m-%Y')
            res.hora = hora
            res.fonoaudiologo = doctor
            res.nombrePaciente = nombre
            res.apellidoPaciente = apellido
            res.rutPaciente = rut
            res.telefonoPaciente = telefono
            res.emailPaciente = email
            res.save()
            
            log = Log(username = email, texto = 'Reserva de Hora')
            log.save()
            
            try:
                subject = 'Confirmación Reserva De Hora - COFAM'
                html_message = render_to_string('reservaHoras/confirmacionCorreo.html', {'fecha': fecha_reserva, 'hora': hora, 'fonoaudiologo': doctor, 'nombre': nombre, 'apellido': apellido, 'rut': rut})
                send_mail(subject, None, settings.EMAIL_HOST_USER, [email], html_message=html_message)
                
            except Exception as e:
                messages.error(request, 'Error al enviar el correo')

            messages.success(request, 'Reserva realizada con éxito')
            return redirect('home')
        else:
            messages.error(request, 'Error al realizar la reserva')
            data["form"] = formulario
    
    return render(request, 'reservaHoras/reservaHora.html', data)

#Cancelar Reserva
@login_required
def cancelarReserva(request, id):
    reserva = ReservaHora.objects.get(id=id)
    fecha_reserva = reserva.fecha.strftime('%d-%m-%Y')
    hora = reserva.hora
    doctor = reserva.fonoaudiologo
    nombre = reserva.nombrePaciente
    apellido = reserva.apellidoPaciente
    rut = reserva.rutPaciente
    email = reserva.emailPaciente
    
    try:
        subject = 'Cancelación Reserva De Hora - COFAM'
        html_message = render_to_string('reservaHoras/cancelarCorreo.html', {'fecha': fecha_reserva, 'hora': hora, 'fonoaudiologo': doctor, 'nombre': nombre, 'apellido': apellido, 'rut': rut})
        send_mail(subject, None, settings.EMAIL_HOST_USER, [email], html_message=html_message)
        reserva.delete()
        
        log = Log(username = email, texto = 'Cancelación de Reserva')
        log.save()
    
        messages.success(request, 'Reserva cancelada con éxito')
                
    except Exception as e:
        messages.error(request, 'Error al cancelar la hora')
    
    return redirect('perfil')

# ------------------- FIN Reserva de horas -------------------

#Restablecer Contrseña
def restablecerContrasena(request, id, token):
    if request.method == 'POST':
        password = request.POST.get('password')
        password2 = request.POST.get('password2')
        
        if password == password2:
            user = User.objects.get(id=id)
            if default_token_generator.check_token(user, token):
                user.set_password(password)
                user.save()
                
                log = Log(username = user.email, texto = 'Restablecimiento de Contraseña')
                log.save()
                
                messages.success(request, 'Contraseña restablecida con éxito')
                return redirect('login')
            else:
                messages.error(request, 'Token inválido')
        else:
            messages.error(request, 'Las contraseñas no coinciden')
    
    return render(request, 'registration/restablecerContrasena.html')

#Reset Contraseña Home
def resetearContrasena(request):
    if request.method == 'POST':
        email = request.POST.get('email')
        if User.objects.filter(email=email).exists():
            user = User.objects.get(email=email)
            try:
                token = default_token_generator.make_token(user)
                reset_url = reverse('restablecerContrasena', args=[user.id, token])
                subject = 'Restablecer Contraseña - COFAM'
                link = request.build_absolute_uri(reset_url)
                html_message = render_to_string('registration/reset_password_email.html', {'reset_url': link})

                send_mail(subject, None, settings.EMAIL_HOST_USER, [email], html_message=html_message)
                
                log = Log(username = email, texto = 'Solicitud de Restablecimiento de Contraseña')
                log.save()
                
                messages.success(request, 'Correo enviado con éxito.')
            except Exception as e:
                messages.error(request, 'Error al enviar el correo')
            return HttpResponseRedirect(request.META.get('HTTP_REFERER'))
        else:
            messages.error(request, 'El correo no está registrado')

    return render(request, 'registration/resetearContrasena.html')

# Registro Fonoaudiologos
@login_required
def registroFono(request):
    data = {
        'form': RegistroFonoForm()
    }
    
    if request.method == 'POST':
        formulario = RegistroFonoForm(request.POST)
        if formulario.is_valid():
            nombre = formulario.cleaned_data.get('nombre')
            apellido = formulario.cleaned_data.get('apellido')
            correo = formulario.cleaned_data.get('email')
           
            if User.objects.filter(email=correo).exists():
                messages.error(request, "El correo ya está registrado.")
            else:
                # Crear Fono
                nuevo_fono = formulario.save()
    
                # Crear un nuevo usuario
                usu = User()
                usu.username = correo
                usu.email = correo
                usu.nombre = nombre
                usu.apellido = apellido
                tipo_usuario_fonoaudiologo = tipo_usuario.objects.get(nombre_tipo_usuario='Fonoaudiologo')
                usu.tipoUsuario = tipo_usuario_fonoaudiologo
                usu.save()

                # Asignar horario de trabajo de lunes a viernes de 09:00 a 18:00
                horario_inicio = time(5, 0)
                horario_fin = time(14, 0)
                dias_semana = [HorasTrabajo.LUNES, HorasTrabajo.MARTES, HorasTrabajo.MIERCOLES, HorasTrabajo.JUEVES, HorasTrabajo.VIERNES]

                for dia in dias_semana:
                    HorasTrabajo.objects.create(
                        doctor=nuevo_fono,
                        dia_semana=dia,
                        hora_inicio=datetime.combine(datetime.today(), horario_inicio),
                        hora_fin=datetime.combine(datetime.today(), horario_fin)
                    )
                
                # Enviar correo de bienvenida y restablecimiento de contraseña
                token = default_token_generator.make_token(usu)
                reset_url = reverse('setPassword', args=[usu.id, token])
                link = request.build_absolute_uri(reset_url)
                subject = 'Bienvenido a COFAM - Configura tu Contraseña'
                html_message = render_to_string('registration/correoBienvenida.html', {'nombre': nombre, 'link': link})
                send_mail(subject, None, settings.EMAIL_HOST_USER, [correo], html_message=html_message)
                
                log = Log(username = correo, texto = 'Registro de Fonoaudiologo')
                log.save()
                
                messages.success(request, f'Fonoaudiologo {nombre} creado. Se ha enviado un correo para configurar la contraseña.')
                return redirect('listaFonos')
            
        else:
            data["form"] = formulario
    
    return render(request, 'registration/registroFono.html', data)

#Comunas
def obtener_comunas(request):
    region_id = request.GET.get('region_id')
    comunas = Comuna.objects.filter(region_id=region_id).values('id', 'comuna').order_by('comuna')
    return JsonResponse(list(comunas), safe=False)


# ------------------- Formularios Evaluación -------------------

#Vista Preguntas
@login_required
def preguntas(request):
    Preguntas = PreguntaFormulario.objects.all()
    data = {
        "preguntas": Preguntas,
    }
    return render(request, 'formularios/preguntas.html', data)

#Crear preguntas
@login_required
def crearPreguntas(request):
    data = {"form": PreguntasForm()}

    if request.method == "POST":
        formulario = PreguntasForm(data=request.POST)
        if formulario.is_valid():
            formulario.save()
            messages.success(request, "Pregunta Creada Correctamente")
            
            log = Log(username = request.user.email, texto = 'Creación de Pregunta')
            log.save()
            
            return redirect(to="preguntas")
        else:
            data["form"] = formulario
    return render(request, 'formularios/crearPreguntas.html', data)

#Modificar preguntas
login_required
def modificarPreguntas(request,id):
    Preguntas = PreguntaFormulario.objects.get(id=id)
    data = {"form": PreguntasForm(instance=Preguntas)}

    if request.method == "POST":
        formulario = PreguntasForm(
            data=request.POST, instance=Preguntas)
        
        if formulario.is_valid():
            formulario.save()
            
            log = Log(username = request.user.email, texto = 'Modificación de Pregunta')
            log.save()
            
            messages.success(request, "Modificado Correctamente")
            return redirect(to="preguntas")
        data["form"] = formulario
    return render(request, 'formularios/modificarPreguntas.html', data)

#Eliminar preguntas
@login_required
def eliminarPreguntas(request, id):
    Preguntas = PreguntaFormulario.objects.get(id=id)
    Preguntas.delete()
    
    log = Log(username = request.user.email, texto = 'Eliminación de Pregunta')
    log.save()
    
    messages.success(request, "Eliminado Correctamente")
    return redirect(to="preguntas")

# ------------------- FIN Formularios Evaluación -------------------

#Sesion Asistida
@login_required
def sesionAsistida(request, id):
    reserva = ReservaHora.objects.get(id=id)
    reserva.estado = 'Asistida'
    reserva.save()
    
    log = Log(username = request.user.email, texto = 'Sesión Asistida')
    log.save()
    
    messages.success(request, "Guardada Correctamente")
    return redirect(to="perfil")

#Sesión No Asistida
@login_required
def sesionNoAsistida(request, id):
    reserva = ReservaHora.objects.get(id=id)
    reserva.estado = 'No Asistid'
    reserva.save()
    
    log = Log(username = request.user.email, texto = 'Sesión No Asistida')
    log.save()
    
    messages.success(request, "Guardada Correctamente")
    return redirect(to="perfil")

# ------------------- Ficha Clinica -------------------

# Ficha Clinica
@login_required
def fichaClinica(request, id):
    paciente = Paciente.objects.get(id=id)
    sesionFono = SesionTerapeutica.objects.filter(paciente=paciente).order_by('-id')
    notas = NotaPaciente.objects.filter(paciente=paciente).order_by('-fecha')
    data = {
        'paciente': paciente,
        'sesiones': sesionFono,
        'notas': notas,
    }
    
    return render(request, 'atencion/fichaClinica.html', data)

@login_required
def formComunicativo(request, id):
    preguntas = PreguntaFormulario.objects.filter(formulario_id=1)
    
    # Obtener el paciente
    paciente = get_object_or_404(Paciente, id=id)
    respuestas = RespuestaFormulario.objects.filter(paciente=paciente, pregunta__formulario_id=1)
    
    if request.method == 'POST':
        all_answered = True
        for pregunta in preguntas:
            respuesta_value = request.POST.get(f'pregunta_{pregunta.id}')
            if respuesta_value is None:
                all_answered = False
                break
        
        if not all_answered:
            messages.error(request, 'Por favor, responda todas las preguntas.')
        else:
            for pregunta in preguntas:
                respuesta_value = request.POST.get(f'pregunta_{pregunta.id}')
                respuesta, created = RespuestaFormulario.objects.get_or_create(
                    pregunta=pregunta,
                    paciente=paciente,
                    defaults={
                        'respuesta': (respuesta_value == 'si'),
                        'fechaRespuesta': timezone.now()
                    }
                )
                respuesta.respuesta = (respuesta_value == 'si')
                respuesta.fechaRespuesta = timezone.now()
                respuesta.save()
            messages.success(request, 'Respuestas guardadas correctamente.')
            return redirect('fichaClinica', id=paciente.id)
    
    # Si hay respuestas, las mostramos
    if respuestas:
        return render(request, 'formularios/formularioUno.html', {'respuestas': respuestas, 'paciente': paciente})
    
    # Si no hay respuestas, mostramos el formulario
    return render(request, 'formularios/formularioUno.html', {'preguntas': preguntas, 'paciente': paciente})

@login_required
def formSocial(request, id):
    preguntas = PreguntaFormulario.objects.filter(formulario_id=2)
    paciente = get_object_or_404(Paciente, id=id)
    respuestas = RespuestaFormulario.objects.filter(paciente=paciente, pregunta__formulario_id=2)
    
    if request.method == 'POST':
        all_answered = True
        for pregunta in preguntas:
            respuesta_value = request.POST.get(f'pregunta_{pregunta.id}')
            if respuesta_value is None or respuesta_value == "":
                all_answered = False
                break
        
        if not all_answered:
            messages.error(request, 'Por favor, responda todas las preguntas seleccionando una opción válida.')
        else:
            for pregunta in preguntas:
                respuesta_value = request.POST.get(f'pregunta_{pregunta.id}')
                if respuesta_value == 'Otro':
                    respuesta_value = request.POST.get(f'otro_{pregunta.id}', '')

                respuesta, created = RespuestaFormulario.objects.get_or_create(
                    pregunta=pregunta,
                    paciente=paciente,
                    defaults={
                        'respuesta': respuesta_value,
                        'fechaRespuesta': timezone.now()
                    }
                )
                respuesta.respuesta = respuesta_value
                respuesta.fechaRespuesta = timezone.now()
                respuesta.save()
            messages.success(request, 'Respuestas guardadas correctamente.')
            return redirect('fichaClinica', id=paciente.id)
    
    # Si hay respuestas, las mostramos
    if respuestas.exists():
        return render(request, 'formularios/formularioDos.html', {'respuestas': respuestas, 'paciente': paciente})
    
    # Si no hay respuestas, mostramos el formulario
    return render(request, 'formularios/formularioDos.html', {'preguntas': preguntas, 'paciente': paciente})

@login_required
def formLenguaje(request, id):
    preguntas = PreguntaFormulario.objects.filter(formulario_id=3)
    paciente = get_object_or_404(Paciente, id=id)
    respuestas = RespuestaFormulario.objects.filter(paciente=paciente, pregunta__formulario_id=3)
    
    if request.method == 'POST':
        all_answered = True
        for pregunta in preguntas:
            respuesta_value = request.POST.get(f'pregunta_{pregunta.id}')
            if respuesta_value is None:
                all_answered = False
                break

        if not all_answered:
            messages.error(request, 'Por favor, responda todas las preguntas.')
        else:
            for pregunta in preguntas:
                respuesta_value = request.POST.get(f'pregunta_{pregunta.id}')
                observacion_value = request.POST.get(f'observacion_{pregunta.id}', '')
                respuesta_text = respuesta_value
                if observacion_value:
                    respuesta_text += f" - Obs: {observacion_value}"
                
                try:
                    respuesta = RespuestaFormulario.objects.create(
                        pregunta=pregunta,
                        paciente=paciente,
                        respuesta=respuesta_text,
                        fechaRespuesta=timezone.now(),
                    )
                    respuesta.save()
                except IntegrityError:
                    respuesta = RespuestaFormulario.objects.get(pregunta=pregunta, paciente=paciente)
                    respuesta.respuesta = respuesta_text
                    respuesta.fechaRespuesta = timezone.now()
                    respuesta.save()
            messages.success(request, 'Respuestas guardadas correctamente.')
            return redirect('fichaClinica', id=paciente.id)
    
    # Si hay respuestas, las mostramos
    if respuestas.exists():
        return render(request, 'formularios/formularioTres.html', {'respuestas': respuestas, 'preguntas': preguntas, 'paciente': paciente})
    
    # Si no hay respuestas, mostramos el formulario
    return render(request, 'formularios/formularioTres.html', {'preguntas': preguntas, 'paciente': paciente})

# Notas Paciente
@login_required
def notasPaciente(request, id):
    paciente = get_object_or_404(Paciente, id=id)

    if request.method == 'POST':
        form = NotasPacienteForm(request.POST)
        if form.is_valid():
            nota = form.save(commit=False)
            nota.paciente = paciente
            nota.save()
            messages.success(request, 'Nota guardada correctamente.')
            return redirect('fichaClinica', id=paciente.id)
        else:
            messages.error(request, 'Error al guardar la nota.')
    else:
        form = NotasPacienteForm()

    data = {
        'paciente': paciente,
        'form': form,
    }
    return render(request, 'atencion/notasPaciente.html', data)

#Sesión Fonoaudiologica
@login_required
@login_required
def sesionFono(request, id):
    paciente = get_object_or_404(Paciente, id=id)
    fono = Fonoaudiologo.objects.get(email=request.user.email)
    data = {
        'form': SesionForm(),
        'paciente': paciente
    }
    
    if request.method == 'POST':
        form = SesionForm(request.POST)
        if form.is_valid():
            sesion = form.save(commit=False)
            sesion.fonoaudiologo = fono
            sesion.paciente = paciente
            sesion.tutor = paciente.tutor
            sesion.fecha = timezone.now()
            sesion.save()
            
            try:
                subject = 'Detalle de Sesión Fonoaudiológica'
                html_message = render_to_string('atencion/detalleSesionEmail.html', {'sesion': sesion})
                send_mail(subject, None, settings.EMAIL_HOST_USER, [paciente.tutor.emailTutor], html_message=html_message)
    
                messages.success(request, 'Sesión guardada correctamente, el correo ha sido enviado al tutor')
                
            except Exception as e:
                messages.error(request, f'Error al guardar la sesión: {e}')
            
            return redirect('fichaClinica', id=paciente.id)  
        else:
            messages.error(request, 'Debe completar todos los campos')
    
    return render(request, 'atencion/sesion.html', data)

#Detalle Sesion
@login_required
def detalleSesion(request, id):
    sesion = get_object_or_404(SesionTerapeutica, id=id)
    return render(request, 'atencion/detalleSesion.html', {'sesion': sesion})

# ------------------- FIN Ficha Clinica -------------------

#Buscar Paciente
@login_required
def busquedaPaciente(request):
    rut = request.GET.get('rut', '')
    pacientes = Paciente.objects.all()
    
    if rut:
        pacientes = pacientes.filter(rut__icontains=rut)
    
    data = {
        'rut': rut,
        'pacientes': pacientes
    }
    
    return render(request, 'atencion/buscarPaciente.html', data)

def buscar_paciente(request):
    rut = request.GET.get('rut')
    
    if rut:
        try:
            paciente = Paciente.objects.get(rut=rut)
            messages.success(request, 'Paciente encontrado')
            url = reverse('fichaClinica', args=[paciente.id])
            return JsonResponse({'found': True, 'url': url})
        except Paciente.DoesNotExist:
            return JsonResponse({'found': False})
    
    return JsonResponse({'found': False})

#Nueva Contraseña Usuario
def nuevaContrasenia(request, id, token):
    user = User.objects.get(id=id)
    if user is not None and default_token_generator.check_token(user, token):
        if request.method == 'POST':
            password = request.POST.get('password')
            password2 = request.POST.get('password2')
            if password == password2:
                user.set_password(password)
                user.save()
                
                log = Log(username = user.email, texto = 'Asignación de Contraseña')
                log.save()
                
                messages.success(request, 'Tu contraseña ha sido establecida con éxito.')
                return redirect('login')
        else:
            return render(request, 'registration/nuevaContrasenia.html')
    else:
        messages.error(request, 'El enlace de Asignación de contraseña no es válido, posiblemente ha caducado.')

    return render(request, 'registration/nuevaContrasenia.html')

#Registro Paciente-Tutor
@login_required
def registroPacienteTutor(request):
    data = {
        "regiones": Region.objects.all(),
        'formPac': RegistroPacienteForm(),
        'formTut': RegistroTutorForm(),
        'rut': request.GET.get('rut', '')  # Modificamos aquí para que si el rut es None, lo reemplace por una cadena vacía
    }
    
    if request.method == 'POST':
        formularioPaciente = RegistroPacienteForm(request.POST)
        formularioTutor = RegistroTutorForm(request.POST)
        if formularioPaciente.is_valid() and formularioTutor.is_valid():
            nombreTutor = formularioTutor.cleaned_data.get('nombreTutor')
            apellidoTutor = formularioTutor.cleaned_data.get('apellidoTutor')
            correoTutor = formularioTutor.cleaned_data.get('emailTutor')
            
            if User.objects.filter(email=correoTutor).exists():
                messages.error(request, "El correo del tutor ya está registrado.")
            else:
                # Crear Tutor
                formularioTutor.save()
                
                tut = Tutor.objects.get(emailTutor=correoTutor)
                pac = Paciente()
                pac.nombre = formularioPaciente.cleaned_data.get('nombre')
                pac.apellido = formularioPaciente.cleaned_data.get('apellido')
                pac.rut = formularioPaciente.cleaned_data.get('rut')
                pac.fechaNacimiento = formularioPaciente.cleaned_data.get('fechaNacimiento')
                pac.genero = formularioPaciente.cleaned_data.get('genero')
                pac.telefono = formularioPaciente.cleaned_data.get('telefono')
                pac.direccion = formularioPaciente.cleaned_data.get('direccion')
                pac.comuna = formularioPaciente.cleaned_data.get('comuna')
                pac.tutor = tut
                pac.save()
                
                # Crear un nuevo usuario
                usuTut = User()
                usuTut.username = correoTutor
                usuTut.email = correoTutor
                usuTut.nombre = nombreTutor
                usuTut.apellido = apellidoTutor
                tipo_usuario_tutor = tipo_usuario.objects.get(nombre_tipo_usuario='Tutor')
                usuTut.tipoUsuario = tipo_usuario_tutor
                usuTut.save()
                
                # Enviar correo de bienvenida y restablecimiento de contraseña
                token = default_token_generator.make_token(usuTut)
                reset_url = reverse('setPassword', args=[usuTut.id, token])
                link = request.build_absolute_uri(reset_url)
                subject = 'Bienvenido a COFAM - Configura tu Contraseña'
                html_message = render_to_string('registration/correoBienvenida.html', {'nombre': nombreTutor, 'link': link})
                send_mail(subject, None, settings.EMAIL_HOST_USER, [correoTutor], html_message=html_message)
                
                log = Log(username=correoTutor, texto='Registro de Paciente y Tutor')
                log.save()
                
                messages.success(request, f'Paciente {pac.nombre} y Tutor {nombreTutor} creados. Se ha enviado un correo para configurar la contraseña.')
                return redirect('fichaClinica', id=pac.id)
            
        else:
            data["formPac"] = formularioPaciente
            data["formTut"] = formularioTutor
    return render(request, 'registration/registroPacienteTutor.html', data)

# Editar Paciente y Tutor
@login_required
def editarPacienteTutor(request, id):
    paciente = get_object_or_404(Paciente, id=id)
    tutor = paciente.tutor

    if request.method == 'POST':
        formPac = RegistroPacienteForm(request.POST, instance=paciente)
        formTut = RegistroTutorForm(request.POST, instance=tutor)
        if formPac.is_valid() and formTut.is_valid():
            formPac.save()
            formTut.save()
            log = Log(username=tutor.emailTutor, texto='Edición de Paciente y Tutor')
            log.save()
            messages.success(request, 'Datos actualizados correctamente.')
            return redirect('fichaClinica', id=paciente.id)
        else:
            messages.error(request, 'Error al actualizar los datos.')
    else:
        formPac = RegistroPacienteForm(instance=paciente)
        formTut = RegistroTutorForm(instance=tutor)

    return render(request, 'registration/editarPacienteTutor.html', {
        'formPac': formPac,
        'formTut': formTut,
        'regiones': Region.objects.all(),
        'paciente': paciente,
        'paciente_region_id': paciente.comuna.region.id if paciente.comuna else None,
        'paciente_comuna_id': paciente.comuna.id if paciente.comuna else None
    })



#Listar Fonos
@login_required
def listarFonos(request):
    fonos = Fonoaudiologo.objects.all()
    data = {
        'fonos': fonos
    }
    return render(request, 'registration/fonoaudiologos.html', data)

#Editar Fono
@login_required
def editarFono(request, id):
    fono = get_object_or_404(Fonoaudiologo, id=id)
    if request.method == "POST":
        form = RegistroFonoForm(request.POST, instance=fono)
        if form.is_valid():
            form.save()
            
            log = Log(username = fono.email, texto = 'Edición de Fonoaudiologo')
            log.save()
            
            messages.success(request, 'Fonoaudiologo actualizado correctamente.')
            return redirect('listaFonos')
    else:
        form = RegistroFonoForm(instance=fono)
    return render(request, 'registration/editarFono.html', {'form': form})

#Eliminar Fono
@login_required
def eliminar_fono(request, id):
    fono = Fonoaudiologo.objects.get(pk=id)
    usuario = User.objects.get(email=fono.email)
    usuario.delete()
    fono.delete()
    
    log = Log(username = fono.email, texto = 'Eliminación de Fonoaudiologo')
    log.save()
    
    messages.success(request, 'Fonoaudiologo eliminado correctamente.')
    return redirect('listaFonos')
    
# ------------------- Reportes -------------------
@login_required
def reportePrincipal(request):
    if request.user.tipoUsuario.nombre_tipo_usuario == 'Gerencia':
        return render(request, 'reportes/reportesPrincipal.html')
    else:
        messages.error(request, 'No tienes permisos para acceder a esta página.')
        return redirect('perfil')
    
@login_required
def export_data_to_excel(request):
    # Crear un libro de trabajo y hojas
    workbook = openpyxl.Workbook()
    fonoaudiologo_sheet = workbook.active
    fonoaudiologo_sheet.title = 'Fonoaudiologos'
    paciente_sheet = workbook.create_sheet(title='Pacientes')
    reserva_sheet = workbook.create_sheet(title='Reservas')
    log_sheet = workbook.create_sheet(title='Logs')

    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    alignment = Alignment(horizontal='center', vertical='center')

    # Encabezados de las hojas
    headers = [
        ('Fonoaudiologos', ['ID', 'Nombre', 'Apellido', 'RUT', 'Género', 'Teléfono', 'Email', 'Clínica']),
        ('Pacientes', ['ID', 'Nombre', 'Apellido', 'RUT', 'Fecha Nacimiento', 'Género', 'Teléfono', 'Dirección', 'Comuna', 'Tutor']),
        ('Reservas', ['ID', 'Fecha', 'Hora', 'Fonoaudiologo', 'Nombre Paciente', 'Apellido Paciente', 'RUT Paciente', 'Teléfono Paciente', 'Email Paciente', 'Estado']),
        ('Logs', ['ID', 'Usuario', 'Fecha', 'Texto'])
    ]
    sheets = [fonoaudiologo_sheet, paciente_sheet, reserva_sheet, log_sheet]

    for sheet, header in zip(sheets, headers):
        sheet.append(header[1])
        for col in range(1, len(header[1]) + 1):
            cell = sheet.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = alignment

    # Ajustar el ancho de las columnas
    column_widths = {
        'Fonoaudiologos': [5, 20, 20, 12, 10, 12, 25, 20],
        'Pacientes': [5, 20, 20, 15, 15, 10, 12, 30, 20, 20],
        'Reservas': [5, 12, 10, 20, 20, 20, 15, 12, 25, 10],
        'Logs': [5, 40, 30, 50]
    }

    for sheet, widths in zip(sheets, column_widths.values()):
        for i, width in enumerate(widths, start=1):
            sheet.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    # Rellenar las hojas con datos
    for fono in Fonoaudiologo.objects.all():
        row = [
            fono.id,
            fono.nombre,
            fono.apellido,
            fono.rut,
            fono.genero.genero,
            fono.telefono,
            fono.email,
            fono.clinica.nombre
        ]
        fonoaudiologo_sheet.append(row)
        for col in range(1, len(row) + 1):
            cell = fonoaudiologo_sheet.cell(row=fonoaudiologo_sheet.max_row, column=col)
            cell.border = thin_border
            cell.alignment = alignment

    for paciente in Paciente.objects.all():
        row = [
            paciente.id,
            paciente.nombre,
            paciente.apellido,
            paciente.rut,
            paciente.fechaNacimiento.strftime('%Y-%m-%d'),  # Formatear fecha
            paciente.genero.genero,
            paciente.telefono,
            paciente.direccion,
            paciente.comuna.comuna,
            paciente.tutor.nombreTutor
        ]
        paciente_sheet.append(row)
        for col in range(1, len(row) + 1):
            cell = paciente_sheet.cell(row=paciente_sheet.max_row, column=col)
            cell.border = thin_border
            cell.alignment = alignment

    for reserva in ReservaHora.objects.all():
        row = [
            reserva.id,
            reserva.fecha.strftime('%Y-%m-%d'),
            reserva.hora.strftime('%H:%M:%S'),
            str(reserva.fonoaudiologo),
            reserva.nombrePaciente,
            reserva.apellidoPaciente,
            reserva.rutPaciente,
            reserva.telefonoPaciente,
            reserva.emailPaciente,
            reserva.estado
        ]
        reserva_sheet.append(row)
        for col in range(1, len(row) + 1):
            cell = reserva_sheet.cell(row=reserva_sheet.max_row, column=col)
            cell.border = thin_border
            cell.alignment = alignment

    # Rellenar la hoja de logs
    for log in Log.objects.order_by('-fecha_inicio'):
        row = [
            log.id,
            log.username,
            log.fecha_inicio.strftime('%Y-%m-%d %H:%M:%S'),  # Formatear fecha y hora
            log.texto
        ]
        log_sheet.append(row)
        for col in range(1, len(row) + 1):
            cell = log_sheet.cell(row=log_sheet.max_row, column=col)
            cell.border = thin_border
            cell.alignment = alignment

    fecha_hoy = datetime.now().strftime('%d-%m-%Y')
    nombre_archivo = f'BaseDeDatos_{fecha_hoy}.xlsx'

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={nombre_archivo}'
    messages.success(request, 'Datos exportados correctamente.')
    workbook.save(response)

    return response

#Reporte Reservas
@login_required
def reporteReservas(request):
    hoy = datetime.now().date().isoformat()
    return render(request, 'reportes/reporteReservas.html', {'hoy': hoy})

@login_required
def filtrar_reservas(request):
    fecha_inicio = request.GET.get('fechaInicio')
    fecha_fin = request.GET.get('fechaFin')
    hoy = datetime.now().date()

    if fecha_inicio and fecha_fin:
        fecha_inicio = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
        fecha_fin = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
        
        reservas = ReservaHora.objects.filter(fecha__range=(fecha_inicio, fecha_fin))
    else:
        reservas = ReservaHora.objects.all()

    context = {
        'reservas': reservas
    }

    html = render_to_string('reportes/tablaReservas.html', context)
    return JsonResponse({'html': html})

@login_required
def exportar_reservas_pdf(request):
    fecha_inicio = request.GET.get('fechaInicio')
    fecha_fin = request.GET.get('fechaFin')

    if fecha_inicio and fecha_fin:
        fecha_inicio = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
        fecha_fin = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
        reservas = ReservaHora.objects.filter(fecha__range=(fecha_inicio, fecha_fin)).order_by('fecha', 'hora')
    else:
        reservas = ReservaHora.objects.all().order_by('fecha', 'hora')

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="reporte_reservas.pdf"'

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), title="Reporte de Reservas")
    elements = []

    styles = getSampleStyleSheet()
    title_style = styles['Title']
    title_style.alignment = 1 
    elements.append(Paragraph("Reporte de Reservas", title_style))

    subtitle_style = styles['Normal']
    elements.append(Paragraph(f"Desde: {fecha_inicio.strftime('%d-%m-%Y')} Hasta: {fecha_fin.strftime('%d-%m-%Y')}", subtitle_style))

    data = [["Fonoaudiologo", "Nombre Paciente", "RUT Paciente", "Teléfono Paciente", "Email Paciente", "Fecha", "Hora"]]

    for reserva in reservas:
        data.append([
            reserva.fonoaudiologo,
            reserva.nombrePaciente + " " + reserva.apellidoPaciente,
            reserva.rutPaciente,
            reserva.telefonoPaciente,
            reserva.emailPaciente,
            reserva.fecha.strftime('%d-%m-%Y'),
            reserva.hora.strftime('%H:%M'),
        ])

    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    elements.append(table)

    doc.build(elements, onFirstPage=lambda canvas, doc: _add_page_number(canvas, doc, "Reporte de Reservas"), 
              onLaterPages=lambda canvas, doc: _add_page_number(canvas, doc, "Reporte de Reservas"))
    pdf = buffer.getvalue()
    buffer.close()
    response.write(pdf)

    return response

def _add_page_number(canvas, doc, title):
    page_num = canvas.getPageNumber()
    text = f"{title} - Página {page_num}"
    canvas.drawRightString(200*mm, 20*mm, text)
    
#Graficos
@login_required
def graficos(request):
    if request.user.tipoUsuario.nombre_tipo_usuario == 'Gerencia':
        # Datos para el gráfico de género
        generos = Genero.objects.all()
        datos_genero = {
            genero.genero: Paciente.objects.filter(genero=genero).count() for genero in generos
        }

        # Datos para el gráfico de rango etario
        pacientes = Paciente.objects.all()
        hoy = date.today()
        edades = [(hoy.year - paciente.fechaNacimiento.year) - ((hoy.month, hoy.day) < (paciente.fechaNacimiento.month, paciente.fechaNacimiento.day)) for paciente in pacientes]
        
        rango_etario = {
            '0-4': 0, '5-9': 0, '10-14': 0, '15-19': 0, '20-24': 0,
            '25-29': 0, '30-34': 0, '35-39': 0, '40-44': 0, '45-49': 0,
            '50-54': 0, '55-59': 0, '60-64': 0, '65-69': 0, '70-74': 0,
            '75-79': 0, '80-84': 0, '85-89': 0, '90-94': 0, '95-99': 0, '100+': 0
        }

        for edad in edades:
            if edad >= 100:
                rango_etario['100+'] += 1
            else:
                key = f"{(edad // 5) * 5}-{(edad // 5) * 5 + 4}"
                rango_etario[key] += 1

        # Datos para el gráfico de pacientes por región
        regiones = Region.objects.all()
        datos_region = {
            region.region: Paciente.objects.filter(comuna__region=region).count() for region in regiones
        }

        # Datos para el gráfico de reservas "Asistida" por fonoaudiólogo
        fonoaudiologos = Fonoaudiologo.objects.all()
        datos_reservas_asistidas = {
            f"{fonoaudiologo.nombre} {fonoaudiologo.apellido}": ReservaHora.objects.filter(fonoaudiologo=fonoaudiologo, estado='Asistida').count() for fonoaudiologo in fonoaudiologos
        }

        context = {
            'datos_genero': datos_genero,
            'rango_etario': rango_etario,
            'datos_region': datos_region,
            'datos_reservas_asistidas': datos_reservas_asistidas
        }
        return render(request, 'reportes/graficos.html', context)
    else:
        messages.error(request, 'No tienes permisos para acceder a esta página.')
        return redirect('perfil')
    
# ------------------- FIN Reportes -------------------
    
#Solicitud OIRS
def modificarOirs(request, solicitud_id):
    solicitud = get_object_or_404(OIRS, id=solicitud_id)
    if request.method == 'POST':
        form = OIRSRespuestaForm(request.POST, instance=solicitud)
        if form.is_valid():
            solicitud = form.save(commit=False)
            solicitud.estado = 'Respondido'
            solicitud.fecha_respuesta = timezone.now()
            solicitud.save()

            # Enviar correo de respuesta
            respuesta = form.cleaned_data['respuesta']
            subject = 'Respuesta a su solicitud OIRS - COFAM'
            from_email = settings.EMAIL_HOST_USER
            recipient_list = [solicitud.email]
            html_message = render_to_string('oirs/respuestaCorreo.html', {
                'solicitud': solicitud,
                'respuesta': respuesta
            })
            
            log = Log(username = solicitud.email, texto = 'Respuesta a Solicitud OIRS - Nro. ' + str(solicitud.id))

            try:
                send_mail(subject, '', from_email, recipient_list, html_message=html_message)
                messages.success(request, 'Respuesta enviada correctamente y correo electrónico enviado.')
            except Exception as e:
                messages.error(request, 'Respuesta enviada, pero hubo un error al enviar el correo electrónico.')

            return redirect('oirs')
    else:
        form = OIRSRespuestaForm(instance=solicitud)
    return render(request, 'oirs/modificarOirs.html', {'solicitud': solicitud, 'form': form})
    
# ------------------- Contacto Tutor - Fono -------------------   

#Primer Contacto 
@login_required
def enviarMensaje(request):
    if request.user.tipoUsuario.nombre_tipo_usuario == 'Tutor':
        fonoaudiologos = Fonoaudiologo.objects.all()
        if request.method == "POST":
            form = MensajeForm(request.POST)
            if form.is_valid():
                mensaje = form.save(commit=False)
                user = request.user.email
                tutor = Tutor.objects.get(emailTutor=user)
                mensaje.paciente = Paciente.objects.get(tutor=tutor)
                fonoaudiologo_id = request.POST.get('fonoaudiologo')
                fonoaudiologo = Fonoaudiologo.objects.get(pk=fonoaudiologo_id)
                mensaje.leidoUno = True
                mensaje.emisor = tutor
                mensaje.receptor = fonoaudiologo
                mensaje.save()
                messages.success(request, 'Mensaje enviado correctamente.')
                return redirect('buzonMensajes')
        else:
            form = MensajeForm()
        return render(request, 'atencion/enviarMensaje.html', {'form': form, 'fonoaudiologos': fonoaudiologos})

#Buzon Mensajes
@login_required
def buzonMensajes(request):

    mensajes_emisor = Mensaje.objects.filter(emisor=request.user.nombre + ' ' + request.user.apellido)
    mensajes_receptor = Mensaje.objects.filter(receptor=request.user.nombre + ' ' + request.user.apellido)
    mensajes = mensajes_emisor | mensajes_receptor
    
    conversaciones = {}
    for mensaje in mensajes:
        destinatario = mensaje.receptor if mensaje.emisor == request.user.nombre + ' ' + request.user.apellido else mensaje.emisor
        if destinatario not in conversaciones or mensaje.fechaEnvio > conversaciones[destinatario].fechaEnvio:
            conversaciones[destinatario] = mensaje
    
    conversaciones_sorted = sorted(conversaciones.values(), key=lambda x: x.fechaEnvio, reverse=True)
    
    return render(request, 'atencion/buzonMensajes.html', {'mensajes': conversaciones_sorted})

#Leer Mensaje
@login_required
def leerMensaje(request, mensajeId):
    if request.user.tipoUsuario.nombre_tipo_usuario == 'Tutor':
        mensaje = get_object_or_404(Mensaje, id=mensajeId)
        fonoaudiologo = mensaje.receptor
        tutor = mensaje.emisor
        mensaje.leidoUno = True
        mensaje.save()
        
    elif request.user.tipoUsuario.nombre_tipo_usuario == 'Fonoaudiologo':
        mensaje = get_object_or_404(Mensaje, id=mensajeId)
        mensaje.leidoDos = True
        mensaje.save()
        fonoaudiologo = mensaje.receptor
        tutor = mensaje.emisor
        

    # Obtener todos los mensajes entre el tutor y el fonoaudiólogo
    mensajes = Mensaje.objects.filter(Q(emisor=tutor, receptor=fonoaudiologo) | Q(emisor=fonoaudiologo, receptor=tutor)).order_by('-fechaEnvio')
    return render(request, 'atencion/leerMensaje.html', {'mensaje': mensaje, 'mensajes': mensajes})

#Responder Mensaje
@login_required
def responderMensaje(request, mensajeId):
    mensaje = get_object_or_404(Mensaje, id=mensajeId)

    if request.method == 'POST':
        form = MensajeForm(request.POST)
        if form.is_valid():
            respuesta = form.save(commit=False)
            respuesta.emisor = request.user.nombre + ' ' + request.user.apellido
            respuesta.receptor = mensaje.emisor if mensaje.emisor != (request.user.nombre + ' ' + request.user.apellido) else mensaje.receptor
            respuesta.paciente = mensaje.paciente

            if request.user.tipoUsuario.nombre_tipo_usuario == 'Tutor':
                respuesta.leidoUno = True
            elif request.user.tipoUsuario.nombre_tipo_usuario == 'Fonoaudiologo':
                respuesta.leidoDos = True

            respuesta.save()
            messages.success(request, 'Mensaje enviado correctamente.')
            return redirect('buzonMensajes')
    else:
        form = MensajeForm()
    
    return render(request, 'atencion/responderMensaje.html', {'form': form, 'mensaje': mensaje})

# ------------------- FIN Contacto Tutor - Fono -------------------   